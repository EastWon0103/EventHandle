from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
#단과대학(객관식), 학과, 동아리, 학번, 이름, 연락처
while True:
    survey = input("조사한 설문지 원본(나가기=q): ")
    if survey == "q":
        break
    try:
        load_wb = load_workbook(survey, data_only=True) 
        load_ws = load_wb['설문지 응답 시트1']
        break
    except:
        print("경로나 파일명 다시 확인하세요, 혹은 시트가 설문지 응답 시트1이 아닐 수도 있습니다.")


inputStatus = input("비교할 저번 이벤트 당첨자 설문지가 있나요?(t/f): ")
if inputStatus == "t":
    while True:
        before_result = input("조사했던 저번 설문지 결과 파일(나가기 q): ")
        if before_result == "q":
            break
        try: 
            load_result_wb = load_workbook(before_result, data_only=True) #!!! 여기 꼭 시작하기전에 수정
            load_result_ws = load_result_wb["Sheet"]
            break
        except:
            print("경로나 파일명 다시 확인하세요, 혹은 시트가 Sheet가 아닐 수도 있습니다.")
else :
    print("중복값 체크 X")

while True:
    try:
        SELECT_SECTION = int(input("선발할 인원 수: "))
        CANDIDATE_SECTION = int(input("후보 인원 수: "))
        break
    except:
        print("----------숫자만 입력하세요------------")

ALL_SECTION = SELECT_SECTION+CANDIDATE_SECTION

#필터링 과정
process_wb = Workbook()
process_ws = process_wb.active

#설문지 결과
result_wb = Workbook()
result_ws = result_wb.active

def makeList(data, ALL_SECTION):
    lst = []
    for i in range(2,ALL_SECTION+2): #102부분을 100명까지 안될 때 수정해야함
        rank = i-1
        college = data.cell(i,2).value
        major = data.cell(i,3).value
        club = data.cell(i,4).value

        try: 
            idNum = str(int(data.cell(i,5).value))
        except ValueError:
            idNum = data.cell(i,5).value

        name = data.cell(i,6).value
        try :
            phone = str(data.cell(i,7).value)
        except :
            phone = str(int(data.cell(i,7).value))
        if (name == None): #아마도 100명이 채 안되었을 경우 인듯? 
            break

        dic = {
            "rank":rank,
            "college":college,  
            "major":major,
            "club":club,
            "idNum":idNum,
            "name":name, 
            "phone":phone}
        lst.append(dic)
    return lst

def xmlForm(xml, SELECT_SECTION):
    xml.column_dimensions["A"].width = 30 #college
    xml.column_dimensions["B"].width = 30 #major
    xml.column_dimensions["C"].width = 18 #club
    xml.column_dimensions["D"].width = 11 #idnumb
    xml.column_dimensions["E"].width = 8 #name
    xml.column_dimensions["F"].width = 16 #phone
    xml.column_dimensions["G"].width = 40 #result
    mergeWidth = 'A'+str(SELECT_SECTION+1)+":"+"G"+str(SELECT_SECTION+1)
    xml.merge_cells(mergeWidth) #예비번호 바꿔야하면 수정

    grayFill = PatternFill(start_color="CCCCCC",
                 end_color="CCCCCC",
                 fill_type="solid")
    xml.cell(SELECT_SECTION+1,1).value = "여기서부터는 예비번호입니다." #예비번호 바꿔야되면 수정
    xml.cell(SELECT_SECTION+1,1).fill = grayFill #여기도 예비번호 바꿔야하면 좌표 수정
    return 

def fillXml(lst, xml,SELECT_SECTION):
    for i in range(len(lst)): #0부터 시작/ 좌표는 1부터 시작이라 +1
        x = i+1  #좌표
        if (i>SELECT_SECTION-1): #50번은 공백(예비번호를 표기) #예비번호가 바뀌면 수정...
            x+=1
        xml.cell(x,1).value = lst[i]["college"]
        xml.cell(x,2).value = lst[i]["major"]
        xml.cell(x,3).value = lst[i]["club"]
        xml.cell(x,4).value = lst[i]["idNum"]
        xml.cell(x,5).value = lst[i]["name"]
        xml.cell(x,6).value = lst[i]["phone"]

def checking_not_overlap(xml, ALL_SECTION, SELECT_SECTION):
    for i in range(1, ALL_SECTION+1): #100명이 아니면 수정...
        if(i==SELECT_SECTION+1): #예비번호 바뀌면 수정...
            continue
        check_ID(xml, i)
        check_phone(xml, i)
        check_club(xml, i)
    twice_overlap(xml,ALL_SECTION,SELECT_SECTION)

def checking(xml, before_xml, ALL_SECTION, SELECT_SECTION): #첫이벤트면 before_xml 빼기
    phoneList = extractionPhone(before_xml) #전명단 폰리스트(첫이벤트면 주석처리)
    for i in range(1, ALL_SECTION+1): #100명이 아니면 수정...
        if(i==SELECT_SECTION+1): #예비번호 바뀌면 수정...
            continue
        check_ID(xml, i)
        check_phone(xml, i)
        check_club(xml, i)
        compareList(xml, i, phoneList) #저번에 중복지원했는지
    twice_overlap(xml, ALL_SECTION, SELECT_SECTION)

def check_ID(xml, i):
    idNum = xml.cell(i,4).value
    if len(idNum)==8 and idNum.isdigit():
        pass
    else:
        try:
            xml.cell(i,7).value+="학번 이상/"
            xml.cell(i,7).fill = orangeFill()
        except TypeError:
            xml.cell(i,7).value="학번 이상/"
            xml.cell(i,7).fill = orangeFill()

#11자리인지 "-"," "를 제외한 문자 (번호)
def check_phone(xml, i):
    str_phone = xml.cell(i,6).value
    result = []
    print(str_phone)
    for w in str_phone:
        if w ==" " or w == "-":
            continue
        else:
            result.append(w)
    format_phone = "".join(result)
    if(len(format_phone))!=11 or (not format_phone.isdigit()):
        try:
            xml.cell(i,7).value+="번호 이상/"
            xml.cell(i,7).fill = orangeFill()
        except TypeError:
            xml.cell(i,7).value="번호 이상/"
            xml.cell(i,7).fill = orangeFill()

#동아리 이상이 있는지(동아리 이름은 노가다로 해야됨....ㅜ.ㅜ)        
def check_club(xml, i):
    club_list = ["b.a.d.a","바다", "bada",
            "busta", "버스타",
            "chorus", "코러스",
            "december", "디셈버",
            "g-chord","gchord","지코드",
            "kookwha", "국화", "국화kookhwa",
            "magenta", "마젠타",
            "musicbox", "뮤직박스", 
            "phil-muse","필뮤즈", "philmuse",
            "themusical","뮤지컬","더뮤지컬",
            "징", "노래모임징",
            "민족극회판갈이", "판갈이",
            "새날",
            "아우성",
            "영상나래",
            "울림패",
            "kusa","쿠사",
            "l.e.o","leo","레오",
            "teatree","티트리",
            "추어오","국민대고양이추어오",
            "꼬마사랑",
            "명운다회",
            "북악기우회", "기우회",
            "손말사랑회", "손말사랑",
            "여행향기", 
            "유스호스텔", 
            "호우회", 
            "lia", "리아",
            "국민서도회", 
            "그림사랑",
            "문예창작회",
            "빛이랑", 
            "c.a.m","cam",
            "ccc", "c.c.c",
            "ivf",
            "기독학생회",
            "네비게이토",
            "불교학생회", "불교동아리",
            "베네딕도",
            "focus","포커스",
            "k.m.t.c","kmtc",
            "kttc",
            "tab", "탭",
            "what'sup", "와썹", "whatsup",
            "국민대산악부", "산악부",
            "북악가오리",
            "북악머슬", 
            "애로우", "에로우","arrow", 
            "윈드밀스", 
            "캐논볼",
            "이카루스",
            "콕",
            "미르택견", "미르",
            "북악검우회", "검우회",
            "유도부",
            "lowland", "로랜드", "로우랜드",
            "태랑",
            "ala","알라",
            "connect","커넥트",
            "ess",
            "kcc",
            "ms",
            "비상구", 
            "우연", "우리역사연구회",
            "창공", "창의공장",
            "프로메테우스",
            "청문회",
            "와이번",
            "레이저 백스", "razorbacks", "레이져 백스","레이저백스","레이져백스"]
    club = xml.cell(i,3).value.lower().replace(" ","") #소문자로 만들어주고/공백없이로 만들어줌
    if not (club in club_list):
        try:
            xml.cell(i,7).value+="동아리 이상/"
            xml.cell(i,7).fill = orangeFill()
        except TypeError:
            xml.cell(i,7).value="동아리 이상/"
            xml.cell(i,7).fill = orangeFill()

#이상있는 부분을 오렌지 색깔로 색칠
def orangeFill():
    orange = PatternFill(start_color="FFCC00",
                    end_color="FFCC00",
                    fill_type="solid")
    return orange 

def compareList(xml, i, phoneList):
    phone = xml.cell(i,6).value
    if phone in phoneList:
        try:
            xml.cell(i,7).value+="전날당첨자/"
            xml.cell(i,7).fill = orangeFill()
        except TypeError:
            xml.cell(i,7).value="전날당첨자/"
            xml.cell(i,7).fill = orangeFill()

def resultForm(xml):
    xml.column_dimensions["A"].width = 30 #college
    xml.column_dimensions["B"].width = 30 #major
    xml.column_dimensions["C"].width = 18 #club
    xml.column_dimensions["D"].width = 11 #idnumb
    xml.column_dimensions["E"].width = 8 #name
    xml.column_dimensions["F"].width = 16 #phone
    xml.column_dimensions["G"].width = 40 #result

def extractionPhone(before_xml):
    phoneList = []
    i=1
    while True:
        if before_xml.cell(i,6).value == None:
            break
        phoneList.append(before_xml.cell(i, 6).value)
        i += 1
    return phoneList

def twice_overlap(xml, ALL_SECTION, SELECT_SECTION):
    lst = []
    key = []
    for i in range(1, ALL_SECTION+2):
        if i == SELECT_SECTION+1:
            continue
        lst.append(xml.cell(i,6).value)
 
    #키값 만들기
    for j in lst:
        if lst.count(j) > 1:
            key.append(j)
    
    realkey = list(set(key))
    print(realkey)
    for k in realkey:
        count = 0
        for x in range(1, ALL_SECTION+2):
            if x == SELECT_SECTION+1:
                print(xml.cell(x,1).value)
                continue
            if xml.cell(x,6).value == k:                
                count+=1
                print(xml.cell(x,6).value, count)
                if count > 1:
                    try:
                        xml.cell(x,7).value+="당일 중복 지원/"
                        xml.cell(x,7).fill = orangeFill()
                    except:
                        xml.cell(x,7).value="당일 중복 지원/"
                        xml.cell(x,7).fill = orangeFill()


#최종결과 만들어주기
def make_result(pc_ws, rs_ws, ALL_SECTION, SELECT_SECTION):
    result = []
    for i in range(1,ALL_SECTION+2): #여기 인원 수정
        if i == SELECT_SECTION+1: #여기도 예비 수정
            continue
        rank = i
        college = pc_ws.cell(i,1).value
        major = pc_ws.cell(i,2).value
        club = pc_ws.cell(i,3).value
        idNum = pc_ws.cell(i,4).value
        name = pc_ws.cell(i,5).value
        phone = pc_ws.cell(i,6).value
        if (pc_ws.cell(i,7).value == None):
            status = True
        else:
            status = False
        
        dic = {
            "rank":rank,
            "college":college,
            "major":major,
            "club":club,
            "idNum":idNum,
            "name":name,
            "phone":phone,
            "status":status
        }
        result.append(dic)
    
    count = 1
    for i in range(len(result)):
        if count > SELECT_SECTION: #여기 몇명 뽑을지, 이상있으면 수정
            break
        if result[i]["status"] == True:
            rs_ws.cell(count,1).value = result[i]["college"]
            rs_ws.cell(count,2).value = result[i]["major"]
            rs_ws.cell(count,3).value = result[i]["club"]
            rs_ws.cell(count,4).value = result[i]["idNum"]
            rs_ws.cell(count,5).value = result[i]["name"]
            rs_ws.cell(count,6).value = result[i]["phone"]
            count += 1
        else:
            continue

cpy = makeList(load_ws, ALL_SECTION)
xmlForm(process_ws, SELECT_SECTION)
fillXml(cpy, process_ws, SELECT_SECTION)
if inputStatus == "t":
    checking(process_ws, load_result_ws, ALL_SECTION, SELECT_SECTION)
else:
    checking_not_overlap(process_ws, ALL_SECTION, SELECT_SECTION)

resultForm(result_ws)
make_result(process_ws, result_ws, ALL_SECTION, SELECT_SECTION)

process_wb.save("process.xlsx")
result_wb.save("result.xlsx")